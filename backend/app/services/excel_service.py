import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Style constants ──────────────────────────────────────────────────────────
DARK_BLUE  = "1F4E79"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ALT_ROW    = "F2F7FB"
WHITE      = "FFFFFF"
GREEN      = "1E7145"

PHASE_COLORS = ["D6E4F0","C6EFCE","FCE4D6","E2EFDA","FFF2CC","DDEBF7","F4CCCC","EAD1DC","D9EAD3"]

_thin  = Side(style="thin",   color="BFBFBF")
_med   = Side(style="medium", color="1F4E79")
BORDER  = Border(left=_thin, right=_thin, top=_thin,  bottom=_thin)
MBORDER = Border(left=_med,  right=_med,  top=_med,   bottom=_med)


# ── Helpers ──────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color)

def _center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _top_left(wrap=True):
    return Alignment(vertical="top", wrap_text=wrap, horizontal="left")

def _header_cell(ws, row, col, value, bg=DARK_BLUE, fg=WHITE, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=fg, size=size)
    c.fill      = _fill(bg)
    c.alignment = _center()
    c.border    = BORDER
    return c

def _data_cell(ws, row, col, value, bold=False, bg=None, align=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=10)
    c.alignment = align or _top_left()
    c.border    = BORDER
    if bg:  c.fill = _fill(bg)
    if fmt: c.number_format = fmt
    return c

def _total_cell(ws, row, col, value, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill      = _fill(GREEN)
    c.alignment = _center(wrap=False)
    c.border    = BORDER
    if fmt: c.number_format = fmt
    return c

def _title_row(ws, title, span, bg=DARK_BLUE, size=13):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=size)
    c.fill      = _fill(bg)
    c.alignment = _center()
    c.border    = MBORDER
    ws.row_dimensions[1].height = 36


# ── Sheet builders ───────────────────────────────────────────────────────────
def _build_cost_breakdown(wb, rows, title):
    ws = wb.create_sheet("Cost Breakdown")
    _title_row(ws, title, span=9, size=13)

    headers    = ["Module","Sub Module","Description","Deliverables",
                  "Assumptions","Frontend Hrs","Backend Hrs","Rate/Hr (₹)","Cost (₹)"]
    col_widths = [24, 26, 46, 42, 42, 15, 14, 13, 14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        _header_cell(ws, 2, ci, h, bg=MED_BLUE)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 32

    DATA_START = 3
    keys = ["module","sub_module","description","deliverables","notes",
            "frontend_hours","backend_hours","rate_per_hour","cost"]

    # Identify module groups for merging
    groups, i = [], 0
    while i < len(rows):
        mod, j = rows[i]["module"], i
        while j < len(rows) and rows[j]["module"] == mod:
            j += 1
        groups.append((mod, i, j))
        i = j

    for ri, row in enumerate(rows, start=DATA_START):
        bg = ALT_ROW if ri % 2 == 0 else None
        for ci, key in enumerate(keys, 1):
            val = row.get(key, "")
            if ci == 1:
                c = ws.cell(row=ri, column=1, value=val)
                c.font      = Font(name="Arial", bold=True, size=10)
                c.fill      = _fill(LIGHT_BLUE)
                c.alignment = _center()
                c.border    = BORDER
            else:
                fmt = "#,##0" if key in ("rate_per_hour", "cost") else None
                _data_cell(ws, ri, ci, val, bg=bg, fmt=fmt)
        ws.row_dimensions[ri].height = 55

    # Merge module column cells per group
    for _, gi, gj in groups:
        sr, er = gi + DATA_START, gj + DATA_START - 1
        if er > sr:
            ws.merge_cells(start_row=sr, start_column=1, end_row=er, end_column=1)
        mc = ws.cell(row=sr, column=1)
        mc.font      = Font(name="Arial", bold=True, size=10)
        mc.fill      = _fill(LIGHT_BLUE)
        mc.alignment = _center()
        mc.border    = BORDER

    # Total row
    tr = len(rows) + DATA_START
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=5)
    _total_cell(ws, tr, 1, "TOTAL")
    _total_cell(ws, tr, 8, "—")

    for col_idx in (6, 7, 9):
        col_letter = get_column_letter(col_idx)
        fmt = "#,##0" if col_idx == 9 else None
        _total_cell(ws, tr, col_idx,
                    f"=SUM({col_letter}{DATA_START}:{col_letter}{tr-1})", fmt=fmt)
    ws.row_dimensions[tr].height = 26
    ws.freeze_panes = "A3"


def _build_cost_summary(wb, rows, title):
    ws = wb.create_sheet("Cost Summary")
    _title_row(ws, f"Cost Summary – {title}", span=3)

    for ci, h in enumerate(["#", "Category", "Cost (₹)"], 1):
        _header_cell(ws, 2, ci, h, bg=MED_BLUE)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 20
    ws.row_dimensions[2].height = 28

    normal = [r for r in rows if r["category"].lower() != "total"]

    for ri, row in enumerate(normal, start=3):
        bg = ALT_ROW if ri % 2 == 0 else None
        _data_cell(ws, ri, 1, ri - 2, bg=bg, align=_center(wrap=False))
        _data_cell(ws, ri, 2, row["category"], bg=bg)
        _data_cell(ws, ri, 3, row["cost"], bg=bg, fmt="#,##0", align=_center(wrap=False))
        ws.row_dimensions[ri].height = 22

    tr = len(normal) + 3
    _total_cell(ws, tr, 1, "")
    _total_cell(ws, tr, 2, "TOTAL")
    _total_cell(ws, tr, 3, f"=SUM(C3:C{tr-1})", fmt="#,##0")
    ws.row_dimensions[tr].height = 26
    ws.freeze_panes = "A3"


def _build_timeline(wb, rows, title):
    ws = wb.create_sheet("Timeline")
    _title_row(ws, f"Project Timeline – {title}", span=3)

    for ci, h in enumerate(["#", "Phase", "Duration"], 1):
        _header_cell(ws, 2, ci, h, bg=MED_BLUE)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 18
    ws.row_dimensions[2].height = 28

    normal = [r for r in rows if "total" not in r["phase"].lower()]
    totals = [r for r in rows if "total"     in r["phase"].lower()]

    for ri, row in enumerate(normal, start=3):
        bg = PHASE_COLORS[(ri - 3) % len(PHASE_COLORS)]
        _data_cell(ws, ri, 1, ri - 2, bg=bg, align=_center(wrap=False))
        _data_cell(ws, ri, 2, row["phase"], bg=bg)
        _data_cell(ws, ri, 3, row["duration"], bg=bg, align=_center(wrap=False))
        ws.row_dimensions[ri].height = 22

    if totals:
        tr = len(normal) + 3
        _total_cell(ws, tr, 1, "")
        _total_cell(ws, tr, 2, totals[0]["phase"])
        _total_cell(ws, tr, 3, totals[0]["duration"])
        ws.row_dimensions[tr].height = 26

    ws.freeze_panes = "A3"


def _build_summary(wb, summary, title):
    ws = wb.create_sheet("Project Summary")
    _title_row(ws, f"Project Summary – {title}", span=2)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30

    fields = [
        ("Total Cost",                   f"₹ {summary['total_cost']:,}"),
        ("Currency",                      summary["currency"]),
        ("Applicable Taxes",             "Yes" if summary["applicable_taxes"] else "No"),
        ("Estimated Project Duration",    summary["estimated_project_duration"]),
        ("Estimated Development Effort",  summary["estimated_development_effort"]),
    ]
    row_bgs = [LIGHT_BLUE, ALT_ROW, LIGHT_BLUE, ALT_ROW, LIGHT_BLUE]

    for ri, ((label, value), bg) in enumerate(zip(fields, row_bgs), start=2):
        c1 = ws.cell(row=ri, column=1, value=label)
        c1.font = Font(name="Arial", bold=True, size=10)
        c1.fill = _fill(bg); c1.alignment = _top_left(wrap=False); c1.border = BORDER

        c2 = ws.cell(row=ri, column=2, value=value)
        c2.font = Font(name="Arial", size=10)
        c2.fill = _fill(bg); c2.alignment = _center(wrap=False); c2.border = BORDER
        ws.row_dimensions[ri].height = 26


# ── Public function ──────────────────────────────────────────────────────────
def generate_xls(llm_raw_response: str, excel_path: str = "cost_estimation_report.xlsx") -> str:
    """
    Parse the LLM JSON response and write a multi-sheet Excel file.

    Expected JSON shape:
    {
      "title": "...",
      "currency": "INR",
      "tables": {
        "cost_breakdown":   [ { module, sub_module, description, deliverables,
                                notes, frontend_hours, backend_hours,
                                rate_per_hour, cost }, ... ],
        "cost_summary":     [ { category, cost }, ... ],
        "timeline_estimate":[ { phase, duration }, ... ]
      },
      "summary": {
        "total_cost": 0,
        "currency": "INR",
        "applicable_taxes": true,
        "estimated_project_duration": "...",
        "estimated_development_effort": "..."
      }
    }

    Args:
        llm_raw_response: Raw string from LLM (may be wrapped in ```json ... ```).
        excel_path:       Output file path.

    Returns:
        Path to the saved Excel file.
    """
    # Accept either a dict or a raw JSON string (with optional ```json fences)
    if isinstance(llm_raw_response, dict):
        data = llm_raw_response
    else:
        raw = llm_raw_response.strip()
        raw = re.sub(r"^```[a-zA-Z]*\n?", "", raw)
        raw = re.sub(r"```$", "", raw.strip())
        data = json.loads(raw)

    title  = data.get("title", "Cost Estimation")
    tables = data["tables"]

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    _build_cost_breakdown(wb, tables["cost_breakdown"],   title)
    _build_cost_summary  (wb, tables["cost_summary"],     title)
    _build_timeline      (wb, tables["timeline_estimate"], title)
    _build_summary       (wb, data["summary"],             title)

    wb.save(excel_path)
    return excel_path

